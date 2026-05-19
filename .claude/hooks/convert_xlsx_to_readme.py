#!/usr/bin/env python3
"""Convert storyboard file → README.md for a video-edits project.

Looks for storyboard in sources/ in priority order:
  1. script_plan.tsv  (preferred — fast, robust to commas in voice script)
  2. script_plan.csv
  3. script_plan.xlsx (requires openpyxl)

Expects 4 columns: Time | Visual | Voice script | Title.
Idempotent: preserves manually-curated README, only regenerates auto-README
when source is newer.
"""
import sys
import csv
from pathlib import Path

AUTO_MARKER = "Auto-generated from"


def read_tsv_or_csv(path: Path, delimiter: str):
    """Read tab- or comma-separated file, return (header, rows)."""
    with path.open(encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=delimiter)
        all_rows = [
            [c.strip() for c in r[:4]]
            for r in reader
            if any(c.strip() for c in r[:4] if c is not None)
        ]
    if not all_rows:
        return None, []
    return all_rows[0], all_rows[1:]


def read_xlsx(path: Path):
    try:
        import openpyxl
    except ImportError:
        print("ERR: openpyxl missing for xlsx", file=sys.stderr)
        return None, []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    header = None
    for row in ws.iter_rows(values_only=True):
        first4 = [str(c).strip() if c is not None else "" for c in row[:4]]
        if not any(first4):
            continue
        if header is None:
            header = first4
            continue
        rows.append(first4)
    return header, rows


def detect_source(sources_dir: Path):
    """Return (path, format) of preferred storyboard file or (None, None)."""
    for fname, fmt in [
        ("script_plan.tsv", "tsv"),
        ("script_plan.csv", "csv"),
        ("script_plan.xlsx", "xlsx"),
    ]:
        p = sources_dir / fname
        if p.exists():
            return p, fmt
    return None, None


def convert(source_path: Path, fmt: str, readme_path: Path, project_name: str) -> bool:
    if fmt == "tsv":
        header, rows = read_tsv_or_csv(source_path, "\t")
    elif fmt == "csv":
        header, rows = read_tsv_or_csv(source_path, ",")
    else:
        header, rows = read_xlsx(source_path)

    if not rows or not header:
        return False

    lines = [
        f"# {project_name}",
        "",
        f"> {AUTO_MARKER} `sources/{source_path.name}` by session-start hook.",
        "> Edit the source file — README regenerates on next session start.",
        "",
        "## Раскадровка",
        "",
        f"| {' | '.join(header)} |",
        f"| {' | '.join(['---'] * len(header))} |",
    ]
    for r in rows:
        cells = [c.replace("\n", " ").replace("|", "\\|") for c in r]
        lines.append(f"| {' | '.join(cells)} |")

    lines += [
        "",
        "## Исходники",
        "",
        "- `sources/narrator_audio*.ogg` — голос диктора",
        "- `sources/footage/` — видео-исходники",
        "- `sources/photos/` — студийные PNG (продукт на чёрном)",
        "- `sources/photos_lifestyle/` — lifestyle JPEG (продукт в среде, если есть)",
        "- `sources/stock/` — стоки (только контекст, не продукт)",
        "",
        "## Следующие шаги (см. CLAUDE.md → workflow A→B)",
        "",
        "- [ ] A1: voice-маркеры через `silencedetect`",
        "- [ ] A2: раскадровка — таблица выше, подтвердить",
        "- [ ] A2.5: gap-list — каких кадров не хватает в footage",
        "- [ ] A3: SFX-палитра",
        "- [ ] A4: музыка-категория",
        "- [ ] B5-B9: layered render",
        "",
    ]
    readme_path.write_text("\n".join(lines), encoding="utf-8")
    return True


def validate_project(proj: Path) -> list:
    """Return list of warnings (missing fields, missing lessons_learned, etc.)."""
    warnings = []
    readme = proj / "README.md"
    final_video = proj / "v_final.mp4"
    lessons = proj / "lessons_learned.md"

    # Required README fields (manual READMEs should have them; auto-gen omits)
    if readme.exists():
        text = readme.read_text(encoding="utf-8", errors="ignore")
        is_auto = AUTO_MARKER in text
        if not is_auto:
            required = {
                "Mood": ["Mood:", "**Mood:**"],
                "Photo policy": ["Photo policy:", "**Photo policy:**"],
            }
            for field, markers in required.items():
                if not any(m in text for m in markers):
                    warnings.append(f"README missing field: {field}")

    # If v_final.mp4 exists but lessons_learned.md doesn't → reminder
    if final_video.exists() and not lessons.exists():
        warnings.append("v_final.mp4 shipped but lessons_learned.md not filled")

    return warnings


def main():
    projects_root = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("video-edits/projects")
    if not projects_root.is_dir():
        return

    status = []
    project_warnings = []
    for proj in sorted(projects_root.iterdir()):
        if not proj.is_dir():
            continue
        sources_dir = proj / "sources"
        if not sources_dir.is_dir():
            continue
        source_path, fmt = detect_source(sources_dir)
        if source_path is None:
            continue

        readme = proj / "README.md"
        readme_is_auto = (
            readme.exists()
            and AUTO_MARKER in readme.read_text(encoding="utf-8", errors="ignore")
        )
        if not readme.exists():
            should_write = True
        elif readme_is_auto and source_path.stat().st_mtime > readme.stat().st_mtime:
            should_write = True
        else:
            should_write = False

        if should_write:
            try:
                if convert(source_path, fmt, readme, proj.name):
                    status.append(f"{proj.name} ({fmt}→README updated)")
                else:
                    status.append(f"{proj.name} ({fmt} empty?)")
            except Exception as e:
                status.append(f"{proj.name} (ERR: {e})")
        else:
            note = (
                "manual README — preserved"
                if readme.exists() and not readme_is_auto
                else f"up-to-date ({fmt})"
            )
            status.append(f"{proj.name} ({note})")

        # Validate every project regardless of conversion
        warns = validate_project(proj)
        for w in warns:
            project_warnings.append(f"{proj.name}: {w}")

    if status:
        print(f"Active video projects: {', '.join(status)}")
    if project_warnings:
        print("⚠  Project warnings:")
        for w in project_warnings:
            print(f"   • {w}")


if __name__ == "__main__":
    main()
